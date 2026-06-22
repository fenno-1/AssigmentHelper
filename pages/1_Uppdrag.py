import json
import os
import uuid
from datetime import date, datetime
from pathlib import Path

import streamlit as st

st.set_page_config(page_title="Uppdrag", page_icon="📁", layout="wide")

CONSULTANTS = ["Manuel Kandala", "Mia Aspberg", "Magnus Sörin"]
STATUSES = ["Öppen", "Intervju", "Pausad", "Vunnen", "Förlorad", "Avböjd"]

# Short-name → canonical mapping used by the Excel importer so that a row with
# "Manuel" resolves to "Manuel Kandala", etc. Add more aliases here as needed.
CONSULTANT_ALIASES = {
    "manuel": "Manuel Kandala",
    "mia": "Mia Aspberg",
    "magnus": "Magnus Sörin",
}

# Excel column header → assignment field. Used by the import-from-Excel UI.
EXCEL_COLUMN_MAP = {
    "Status": "status",
    "Datum": "date",
    "Konsult": "consultant",
    "Roll": "name",
    "Kund": "customer",
    "Mäklare": "broker",
    "Länk till uppdraget": "url",
    "Kontaktperson": "contact_person",
    "Telefon": "contact_phone",
    "Mail": "contact_email",
    "Lämnat timpris": "price",
}
EXCEL_COMMENT_COL = "Kommentar"
# Path is overridable via env var so deployments (e.g. App Service) can point
# at persistent storage like /home/data/assignments.json. Defaults to the
# repo-relative file for local development.
DATA_FILE = Path(os.environ.get("ASSIGNMENTS_PATH", str(Path(__file__).parent.parent / "assignments.json")))

LIST_COLUMNS = [
    "name",
    "date",
    "consultant",
    "status",
    "customer",
    "broker",
    "url",
    "contact_person",
    "contact_phone",
    "contact_email",
    "price",
]

COLUMN_LABELS = {
    "name": "Uppdragsnamn",
    "date": "Datum",
    "consultant": "Konsult",
    "status": "Status",
    "customer": "Kund",
    "broker": "Förmedlare",
    "url": "URL",
    "contact_person": "Kontaktperson",
    "contact_phone": "Telefon",
    "contact_email": "E-post",
    "price": "Pris (kr/h)",
}


def load_assignments() -> list[dict]:
    if DATA_FILE.exists():
        with open(DATA_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []


def save_assignments(assignments: list[dict]) -> None:
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(assignments, f, ensure_ascii=False, indent=2)


def _normalize_consultant(value) -> str:
    """Map a raw consultant value (e.g. 'Manuel') to the canonical full name.
    Falls back to the first configured consultant for unknown / empty values."""
    if value is None:
        return CONSULTANTS[0]
    key = str(value).strip().lower()
    if not key:
        return CONSULTANTS[0]
    if key in CONSULTANT_ALIASES:
        return CONSULTANT_ALIASES[key]
    for full in CONSULTANTS:
        if key == full.lower():
            return full
    return CONSULTANTS[0]


def _excel_date_to_iso(value) -> str:
    """Coerce an Excel datetime/date/string cell into YYYY-MM-DD."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None or (isinstance(value, str) and not value.strip()):
        return str(date.today())
    return str(value).strip()


def parse_excel(file) -> tuple[list[dict], list[str]]:
    """Read an Excel upload and return (parsed_assignments, warnings).

    Validates that the workbook has the expected Swedish column headers, then
    maps each data row into the same dict shape used elsewhere on this page —
    new UUID id, ISO date, normalized consultant, Kommentar promoted to a
    single timestamped note. Empty / header-only files yield ([], [warning])."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["openpyxl är inte installerat på servern."]

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return [], [f"Kunde inte läsa filen: {exc}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Filen är tom."]

    headers = [h.strip() if isinstance(h, str) else h for h in rows[0]]
    required = list(EXCEL_COLUMN_MAP.keys()) + [EXCEL_COMMENT_COL]
    missing = [c for c in required if c not in headers]
    if missing:
        return [], [f"Saknade kolumner: {', '.join(missing)}"]

    col_idx = {h: i for i, h in enumerate(headers)}
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    parsed: list[dict] = []
    warnings: list[str] = []

    for r_num, row in enumerate(rows[1:], start=2):
        # Skip blank rows so a trailing empty line in the sheet is harmless
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue

        name_val = row[col_idx["Roll"]]
        if not name_val or not str(name_val).strip():
            warnings.append(f"Rad {r_num}: saknar 'Roll' — hoppar över.")
            continue

        status = (row[col_idx["Status"]] or STATUSES[0])
        status = str(status).strip()
        if status not in STATUSES:
            warnings.append(f"Rad {r_num}: okänd status '{status}' → '{STATUSES[0]}'.")
            status = STATUSES[0]

        price_raw = row[col_idx["Lämnat timpris"]]
        if price_raw is None or (isinstance(price_raw, str) and not price_raw.strip()):
            price = None
        else:
            try:
                price = int(price_raw)
            except (TypeError, ValueError):
                warnings.append(f"Rad {r_num}: pris '{price_raw}' kunde inte tolkas → tomt.")
                price = None

        def _text(col: str) -> str:
            v = row[col_idx[col]]
            return str(v).strip() if v is not None else ""

        assignment = {
            "id": str(uuid.uuid4()),
            "name": str(name_val).strip(),
            "date": _excel_date_to_iso(row[col_idx["Datum"]]),
            "consultant": _normalize_consultant(row[col_idx["Konsult"]]),
            "status": status,
            "customer": _text("Kund"),
            "broker": _text("Mäklare"),
            "url": _text("Länk till uppdraget"),
            "contact_person": _text("Kontaktperson"),
            "contact_phone": _text("Telefon"),
            "contact_email": _text("Mail"),
            "price": price,
        }

        comment = row[col_idx[EXCEL_COMMENT_COL]]
        if comment and str(comment).strip():
            assignment["notes"] = [{"text": str(comment).strip(), "timestamp": now}]

        parsed.append(assignment)

    return parsed, warnings


def build_excel_export(assignments: list[dict]) -> bytes:
    """Render assignments to an .xlsx workbook and return its bytes.

    Each assignment is one row using the same columns as the list view. Its
    notes follow as grouped (outline level 1) rows directly beneath it, so
    Excel shows a +/- control to expand/collapse the notes per assignment.
    Returns an empty byte string if openpyxl is unavailable."""
    try:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = "Uppdrag"
    # Summary (assignment) row sits above its note detail rows, so the outline
    # expand button must render on the row above the group, not below it.
    ws.sheet_properties.outlinePr.summaryBelow = False

    headers = [COLUMN_LABELS[k] for k in LIST_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for a in assignments:
        row = []
        for k in LIST_COLUMNS:
            v = a.get(k)
            if k == "url":
                v = _normalize_url(v) if v else ""
            row.append(v if v is not None else "")
        ws.append(row)

        for note in a.get("notes", []):
            ws.append([note.get("timestamp", ""), note.get("text", "")])
            r = ws.max_row
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
            ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Reasonable column widths; notes text (col B) gets the most room.
    widths = {"A": 22, "B": 50, "C": 16, "D": 12, "E": 20, "F": 20,
              "G": 30, "H": 18, "I": 14, "J": 24, "K": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def add_note(assignment_id: str, text: str) -> None:
    assignments = load_assignments()
    for a in assignments:
        if a["id"] == assignment_id:
            if "notes" not in a:
                a["notes"] = []
            a["notes"].append({
                "text": text,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            break
    save_assignments(assignments)


def _normalize_url(url: str) -> str:
    """Ensure a URL has a scheme so it renders as an absolute, clickable link.
    Without this, values like 'www.cinode.com/...' are treated as relative."""
    url = (url or "").strip()
    if url and not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def delete_assignment(assignment_id: str) -> None:
    assignments = load_assignments()
    assignments = [a for a in assignments if a["id"] != assignment_id]
    save_assignments(assignments)


def empty_assignment() -> dict:
    return {
        "id": str(uuid.uuid4()),
        "name": "",
        "date": str(date.today()),
        "consultant": CONSULTANTS[0],
        "status": "Öppen",
        "customer": "",
        "broker": "",
        "url": "",
        "description": "",
        "contact_person": "",
        "contact_phone": "",
        "contact_email": "",
        "price": None,
    }


def assignment_form(assignment: dict, is_new: bool) -> None:
    with st.form("assignment_form"):
        st.subheader("Nytt uppdrag" if is_new else "Redigera uppdrag")

        name = st.text_input("Uppdragsnamn *", value=assignment.get("name", ""))
        date_val = st.text_input("Datum (YYYY-MM-DD)", value=assignment.get("date", str(date.today())))

        consultant_idx = CONSULTANTS.index(assignment["consultant"]) if assignment.get("consultant") in CONSULTANTS else 0
        consultant = st.selectbox("Konsult *", CONSULTANTS, index=consultant_idx)

        status_idx = STATUSES.index(assignment["status"]) if assignment.get("status") in STATUSES else 0
        status = st.selectbox("Status", STATUSES, index=status_idx)

        customer = st.text_input("Kund", value=assignment.get("customer", ""))
        broker = st.text_input("Förmedlare", value=assignment.get("broker", ""))
        url = st.text_input("Uppdragets URL", value=assignment.get("url", ""))
        saved_url = _normalize_url(assignment.get("url", ""))
        if saved_url:
            st.markdown(
                f'🔗 <a href="{saved_url}" target="_blank" rel="noopener noreferrer">Öppna uppdragslänk</a>',
                unsafe_allow_html=True,
            )
        description = st.text_area(
            "Uppdragsbeskrivning",
            value=assignment.get("description", ""),
            height=200,
            help="Fritext. Fyll i manuellt när uppdraget saknar URL.",
        )

        contact_person = st.text_input("Kontaktperson", value=assignment.get("contact_person", ""))
        contact_phone = st.text_input("Telefon", value=assignment.get("contact_phone", ""))
        contact_email = st.text_input("E-post", value=assignment.get("contact_email", ""))

        price_val = assignment.get("price")
        price = st.number_input(
            "Pris (kr/h)",
            min_value=0,
            step=25,
            value=int(price_val) if price_val is not None else 0,
        )

        col_save, col_cancel = st.columns([1, 5])
        submitted = col_save.form_submit_button("Spara", type="primary")
        cancelled = col_cancel.form_submit_button("Avbryt")

    # Delete lives outside the form so it can use a two-step confirmation
    # (a form_submit_button would also save/validate the open fields).
    if not is_new:
        st.divider()
        if not st.session_state.get("confirm_delete"):
            if st.button("🗑 Ta bort", key="delete_assignment"):
                st.session_state["confirm_delete"] = True
                st.rerun()
        else:
            st.warning("Är du säker på att du vill ta bort uppdraget? Det går inte att ångra.")
            col_yes, col_no = st.columns([1, 5])
            if col_yes.button("Ja, ta bort", type="primary", key="confirm_delete_yes"):
                delete_assignment(assignment["id"])
                st.session_state.pop("editing_id", None)
                st.session_state.pop("confirm_delete", None)
                st.success("Uppdraget togs bort.")
                st.rerun()
            if col_no.button("Avbryt", key="confirm_delete_no"):
                st.session_state.pop("confirm_delete", None)
                st.rerun()

    if cancelled:
        st.session_state.pop("confirm_delete", None)
        st.session_state.pop("editing_id", None)
        st.session_state.pop("creating", None)
        st.rerun()

    if submitted:
        if not name.strip():
            st.error("Uppdragsnamn är obligatoriskt.")
            return
        if not consultant:
            st.error("Konsult är obligatoriskt.")
            return

        updated = {
            **assignment,
            "name": name.strip(),
            "date": date_val.strip(),
            "consultant": consultant,
            "status": status,
            "customer": customer.strip(),
            "broker": broker.strip(),
            "url": url.strip(),
            "description": description.strip(),
            "contact_person": contact_person.strip(),
            "contact_phone": contact_phone.strip(),
            "contact_email": contact_email.strip(),
            "price": int(price) if price else None,
        }

        assignments = load_assignments()
        if is_new:
            assignments.append(updated)
        else:
            assignments = [updated if a["id"] == updated["id"] else a for a in assignments]

        save_assignments(assignments)
        st.session_state.pop("editing_id", None)
        st.session_state.pop("creating", None)
        st.session_state.pop("confirm_delete", None)
        st.success("Sparat!")
        st.rerun()


# --- Main ---

st.title("Uppdrag")

if "editing_id" not in st.session_state:
    st.session_state["editing_id"] = None
if "creating" not in st.session_state:
    st.session_state["creating"] = False

# Show form when creating or editing
if st.session_state["creating"]:
    assignment_form(empty_assignment(), is_new=True)
    st.stop()

if st.session_state["editing_id"]:
    assignments = load_assignments()
    match = next((a for a in assignments if a["id"] == st.session_state["editing_id"]), None)
    if match:
        assignment_form(match, is_new=False)
        st.stop()
    else:
        st.session_state["editing_id"] = None

# List view
col_title, col_btn = st.columns([6, 1])
col_btn.button(
    "Nytt uppdrag",
    type="primary",
    on_click=lambda: st.session_state.update({"creating": True}),
)

# --- Excel import ---
# Collapsed expander so it stays out of the way until needed. Importing is
# append-only and lives under the same Entra-gated UI as the rest of the page.
with st.expander("📂 Importera från Excel"):
    st.caption(
        "Förväntade kolumner: " + ", ".join(list(EXCEL_COLUMN_MAP.keys()) + [EXCEL_COMMENT_COL]) + "."
    )
    uploaded = st.file_uploader("Välj .xlsx-fil", type=["xlsx"], key="import_xlsx")
    if uploaded is not None:
        signature = (uploaded.name, uploaded.size)
        if st.session_state.get("last_imported_signature") == signature:
            st.info(
                f"'{uploaded.name}' har redan importerats i denna session. "
                "Ta bort filen ovan för att importera igen."
            )
        else:
            parsed, warnings = parse_excel(uploaded)
            for w in warnings:
                st.warning(w)
            if parsed:
                st.write(f"**Förhandsgranskning – {len(parsed)} uppdrag att importera:**")
                preview = [
                    {COLUMN_LABELS[k]: a[k] for k in LIST_COLUMNS}
                    for a in parsed
                ]
                st.dataframe(preview, use_container_width=True, hide_index=True)
                if st.button(f"Importera {len(parsed)} uppdrag", type="primary", key="confirm_import"):
                    existing = load_assignments()
                    save_assignments(existing + parsed)
                    st.session_state["last_imported_signature"] = signature
                    st.success(f"Importerade {len(parsed)} uppdrag.")
                    st.rerun()

assignments = load_assignments()

if not assignments:
    st.info("Inga uppdrag registrerade ännu. Klicka på 'Nytt uppdrag' för att börja.")
    st.stop()

# Sort state: clicking a column header sorts by it; clicking the same header
# again toggles the direction. Defaults to date descending.
if "sort_key" not in st.session_state:
    st.session_state["sort_key"] = "date"
    st.session_state["sort_desc"] = True


def _sort_value(a: dict):
    v = a.get(st.session_state["sort_key"])
    if st.session_state["sort_key"] == "price":
        # Numeric sort; missing prices sort first ascending / last descending.
        return v if isinstance(v, (int, float)) else float("-inf")
    return str(v if v is not None else "").lower()


assignments_sorted = sorted(
    assignments, key=_sort_value, reverse=st.session_state["sort_desc"]
)

# Build display table with a clickable name column
st.write("Klicka på ett uppdragsnamn för att redigera.")

COL_WIDTHS = [2, 1, 1.5, 1, 1.5, 1.5, 2, 1.5, 1.2, 1.8, 1]
# Columns with a fixed value set get a dropdown filter; status gets a
# multiselect (defaulting to the active statuses); the rest get free-text
# substring filtering.
SELECT_FILTERS = {"consultant": CONSULTANTS}
MULTI_FILTERS = {"status": STATUSES}
STATUS_DEFAULT = ["Öppen", "Intervju"]

# Seed the status multiselect once so we can both render its popover label and
# read the selection without passing default= alongside key= (which warns).
if "filter_status" not in st.session_state:
    st.session_state["filter_status"] = STATUS_DEFAULT

header_cols = st.columns(COL_WIDTHS)
filters: dict[str, object] = {}
for col, key in zip(header_cols, LIST_COLUMNS):
    # Header doubles as a sort toggle: an arrow marks the active sort column and
    # its direction; clicking the active column flips direction.
    arrow = ""
    if st.session_state["sort_key"] == key:
        arrow = " ▼" if st.session_state["sort_desc"] else " ▲"
    if col.button(
        f"{COLUMN_LABELS[key]}{arrow}",
        key=f"sort_{key}",
        use_container_width=True,
    ):
        if st.session_state["sort_key"] == key:
            st.session_state["sort_desc"] = not st.session_state["sort_desc"]
        else:
            st.session_state["sort_key"] = key
            st.session_state["sort_desc"] = False
        st.rerun()
    if key in MULTI_FILTERS:
        # A multiselect is unusable in this narrow column, so put it inside a
        # full-width popover; the trigger button still sits under the header.
        selected = st.session_state.get(f"filter_{key}", [])
        trigger = f"Filter ({len(selected)})" if selected else "Alla"
        with col.popover(trigger, use_container_width=True):
            filters[key] = st.multiselect(
                COLUMN_LABELS[key],
                MULTI_FILTERS[key],
                key=f"filter_{key}",
                label_visibility="collapsed",
            )
    elif key in SELECT_FILTERS:
        filters[key] = col.selectbox(
            COLUMN_LABELS[key],
            ["Alla"] + SELECT_FILTERS[key],
            key=f"filter_{key}",
            label_visibility="collapsed",
        )
    else:
        filters[key] = col.text_input(
            COLUMN_LABELS[key],
            key=f"filter_{key}",
            label_visibility="collapsed",
            placeholder="Filtrera",
        )


def _matches_filters(a: dict) -> bool:
    for key, val in filters.items():
        cell = str(a.get(key) if a.get(key) is not None else "")
        if key in MULTI_FILTERS:
            # Empty selection = no filtering (show all statuses)
            if val and cell not in val:
                return False
        elif key in SELECT_FILTERS:
            if val and val != "Alla" and cell != val:
                return False
        elif val and val.lower() not in cell.lower():
            return False
    return True


filtered = [a for a in assignments_sorted if _matches_filters(a)]

st.divider()

if not filtered:
    st.info("Inga uppdrag matchar filtret.")
    st.stop()

# Export the currently filtered/sorted view. Notes ride along as grouped rows.
st.download_button(
    f"⬇ Exportera {len(filtered)} uppdrag till Excel",
    data=build_excel_export(filtered),
    file_name=f"uppdrag_{date.today().isoformat()}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

for assignment in filtered:
    row_cols = st.columns(COL_WIDTHS)

    # Name as a button to trigger edit
    if row_cols[0].button(assignment.get("name", "(inget namn)"), key=f"edit_{assignment['id']}"):
        st.session_state["editing_id"] = assignment["id"]
        st.rerun()

    row_cols[1].write(assignment.get("date", ""))
    row_cols[2].write(assignment.get("consultant", ""))
    row_cols[3].write(assignment.get("status", ""))
    row_cols[4].write(assignment.get("customer", ""))
    row_cols[5].write(assignment.get("broker", ""))

    url = _normalize_url(assignment.get("url", ""))
    if url:
        row_cols[6].markdown(
            f'<a href="{url}" target="_blank" rel="noopener noreferrer">länk</a>',
            unsafe_allow_html=True,
        )
    else:
        row_cols[6].write("")

    row_cols[7].write(assignment.get("contact_person", ""))
    row_cols[8].write(assignment.get("contact_phone", ""))
    row_cols[9].write(assignment.get("contact_email", ""))

    price = assignment.get("price")
    row_cols[10].write(str(price) if price is not None else "")

    notes = assignment.get("notes", [])
    label = f"Noteringar ({len(notes)})" if notes else "Noteringar"
    with st.expander(label):
        for note in notes:
            st.markdown(f"**{note['timestamp']}**")
            st.write(note["text"])
            st.divider()
        note_key = f"note_input_{assignment['id']}"
        new_note = st.text_area("Ny notering", key=note_key, height=80)
        if st.button("Lägg till notering", key=f"add_note_{assignment['id']}"):
            if new_note.strip():
                add_note(assignment["id"], new_note.strip())
                st.rerun()
