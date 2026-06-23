"""One-off importer for data/historik-sokta-uppdrag.xlsx.

The historical workbook comes from another system, so its column names, status
vocabulary and consultant names don't match what pages/1_Uppdrag.py expects.
This script transforms each row into the same assignment dict shape used by the
app and merges it into an assignments.json file.

Decisions baked in (agreed with the product owner 2026-06-22):
  * Status "Stängd"/"stängd" -> "Förlorad"; "Vunnet" -> "Vunnen"; empty -> "Förlorad".
  * Consultant short names / typos are canonicalised to one name per person;
    multi-person and unknown values ("Flera", "?", "Jack, Illia") are kept as-is.
  * Rows without an assignment name are skipped (same as the Excel UI importer).
  * Kommentar becomes a single note timestamped with the assignment's own date.
  * Each record is tagged source="historik-import" so it can be found/rolled back.

Usage:
    python scripts/import_historik.py --dry-run            # stats only, writes nothing
    python scripts/import_historik.py --out out.json       # write transformed records
    python scripts/import_historik.py --merge assignments.json   # append into existing file
"""
import argparse
import json
import uuid
from collections import Counter
from datetime import date, datetime
from pathlib import Path

XLSX = Path(__file__).parent.parent / "data" / "historik-sokta-uppdrag.xlsx"

# Source column order in Sheet1.
COLS = ["Status", "Datum", "Konsult", "Uppdragsnamn", "Kommentar", "Kund",
        "Mäklare", "URL", "Kontaktperson", "Telefon", "Epost", "Pris (kr/h)"]

STATUS_MAP = {
    "stängd": "Förlorad",
    "vunnet": "Vunnen",
    "vunnen": "Vunnen",
}
EMPTY_STATUS = "Förlorad"

# Raw consultant value (lower-cased, stripped) -> canonical name. Anything not
# listed here is kept verbatim (covers multi-person and unknown values).
CONSULTANT_CANON = {
    "illia": "Illia Shypko",
    "illia shypko": "Illia Shypko",
    "manuel": "Manuel Kandala",
    "mia": "Mia Aspberg",
    "mia aspberg": "Mia Aspberg",
    "jonas": "Jonas Åkerfeldt",
    "jonas åkerfeldt": "Jonas Åkerfeldt",
    "andreas": "Andreas Behrendtz",
    "andreas behrendtz": "Andreas Behrendtz",
    "jack": "Jack Johansson",
    "jack johansson": "Jack Johansson",
    "christoffer": "Christoffer Atle",
    "christofer": "Christoffer Atle",
    "chriostoffer": "Christoffer Atle",
    "christoffer atle": "Christoffer Atle",
}


def _canon_consultant(value) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    return CONSULTANT_CANON.get(raw.lower(), raw)


def _map_status(value) -> str:
    if value is None or not str(value).strip():
        return EMPTY_STATUS
    return STATUS_MAP.get(str(value).strip().lower(), str(value).strip())


def _iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None or not str(value).strip():
        return ""
    return str(value).strip()


def _parse_price(value):
    """Return (price:int|None, leftover_note:str|None).

    '1050' -> 1050; '920?' -> 920; free text like 'Väldigt pressat...' keeps the
    price empty but is preserved as a note so the information isn't lost."""
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return int(value), None
    text = str(value).strip()
    if not text:
        return None, None
    # Only treat as a price if the whole value is a number (allowing a trailing
    # '?' and internal spaces, e.g. '920?' or '1 350'). Free-text price notes
    # like 'Väldigt pressat, runt 800...' keep price empty but become a note.
    cleaned = text.rstrip("?").replace(" ", "")
    if cleaned.isdigit():
        return int(cleaned), None
    return None, f"Pris (från historik): {text}"


def _text(value) -> str:
    return str(value).strip() if value is not None else ""


def transform() -> tuple[list[dict], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {name: i for i, name in enumerate(COLS)}

    out: list[dict] = []
    warnings: list[str] = []
    for r_num, row in enumerate(rows[1:], start=2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        name = _text(row[idx["Uppdragsnamn"]])
        if not name:
            warnings.append(f"Rad {r_num}: saknar uppdragsnamn — hoppar över.")
            continue

        iso = _iso_date(row[idx["Datum"]])
        price, price_note = _parse_price(row[idx["Pris (kr/h)"]])

        a = {
            "id": str(uuid.uuid4()),
            "name": name,
            "date": iso or str(date.today()),
            "consultant": _canon_consultant(row[idx["Konsult"]]),
            "status": _map_status(row[idx["Status"]]),
            "customer": _text(row[idx["Kund"]]),
            "broker": _text(row[idx["Mäklare"]]),
            "url": _text(row[idx["URL"]]),
            "description": "",
            "contact_person": _text(row[idx["Kontaktperson"]]),
            "contact_phone": _text(row[idx["Telefon"]]),
            "contact_email": _text(row[idx["Epost"]]),
            "price": price,
            "source": "historik-import",
        }

        notes = []
        ts = iso or str(date.today())
        comment = row[idx["Kommentar"]]
        if comment is not None and str(comment).strip():
            notes.append({"text": str(comment).strip(), "timestamp": ts})
        if price_note:
            notes.append({"text": price_note, "timestamp": ts})
        if notes:
            a["notes"] = notes

        out.append(a)

    return out, warnings


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", help="Write transformed records to this JSON file.")
    ap.add_argument("--merge", help="Append transformed records into this existing assignments JSON file (a .bak backup is written first).")
    ap.add_argument("--dry-run", action="store_true", help="Print stats only.")
    args = ap.parse_args()

    records, warnings = transform()
    for w in warnings:
        print("WARN:", w)

    print(f"\nTransformerade {len(records)} uppdrag.")
    print("Status:", dict(Counter(a["status"] for a in records)))
    print("Konsulter:", dict(Counter(a["consultant"] or "(tom)" for a in records).most_common()))
    print("Med noteringar:", sum(1 for a in records if a.get("notes")))
    print("Med pris:", sum(1 for a in records if a.get("price") is not None))

    if args.dry_run:
        print("\n--dry-run: inget skrivet.")
        print("Stickprov:", json.dumps(records[0], ensure_ascii=False, indent=2))
        return

    if args.out:
        Path(args.out).write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Skrev {len(records)} poster till {args.out}")

    if args.merge:
        target = Path(args.merge)
        existing = json.loads(target.read_text(encoding="utf-8")) if target.exists() else []
        backup = target.with_suffix(target.suffix + ".bak")
        backup.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Säkerhetskopia: {backup} ({len(existing)} befintliga poster)")
        merged = existing + records
        target.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Skrev {len(merged)} poster ({len(existing)} + {len(records)}) till {target}")


if __name__ == "__main__":
    main()
